import sqlite3
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

# Connexion à la base de données
conn = sqlite3.connect('/root/.openclaw/workspace/agents_immobiliers.db')
cursor = conn.cursor()

# Créer un nouveau classeur Excel
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Agents Immobiliers IDF"

# Styles
header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Bordures
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# En-têtes
headers = [
    "N°", "Nom Agence", "Adresse", "Téléphone", "Email", 
    "Site Web", "Note Google", "Nb Avis", "Zone", "Status Email"
]

# Écrire les en-têtes
for col_num, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_num)
    cell.value = header
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = header_alignment
    cell.border = thin_border

# Récupérer tous les agents
cursor.execute("""
    SELECT name, address, phone, email, website, rating, reviews, zone 
    FROM agents_immobiliers 
    ORDER BY zone, name
""")
agents = cursor.fetchall()

# Écrire les données
row_num = 2
for idx, agent in enumerate(agents, 1):
    name, address, phone, email, website, rating, reviews, zone = agent
    
    # Déterminer le status de l'email
    if email and email != 'N/A' and '@' in email:
        # Vérifier si c'est un email valide (pas une image)
        if any(ext in email.lower() for ext in ['.jpg', '.png', '.gif', '.webp']):
            email_status = "⚠️ À vérifier"
            email = ""
        else:
            email_status = "✅ OK"
    else:
        email_status = "❌ Manquant"
        email = ""
    
    ws.cell(row=row_num, column=1, value=idx)
    ws.cell(row=row_num, column=2, value=name)
    ws.cell(row=row_num, column=3, value=address)
    ws.cell(row=row_num, column=4, value=phone)
    ws.cell(row=row_num, column=5, value=email)
    ws.cell(row=row_num, column=6, value=website)
    ws.cell(row=row_num, column=7, value=rating)
    ws.cell(row=row_num, column=8, value=reviews)
    ws.cell(row=row_num, column=9, value=zone)
    ws.cell(row=row_num, column=10, value=email_status)
    
    # Appliquer les bordures
    for col in range(1, 11):
        ws.cell(row=row_num, column=col).border = thin_border
    
    row_num += 1

# Ajuster la largeur des colonnes
column_widths = [5, 40, 50, 15, 30, 40, 10, 10, 15, 15]
for i, width in enumerate(column_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = width

# Ajouter une ligne de résumé
row_num += 1
ws.cell(row=row_num, column=1, value="RÉSUMÉ")
ws.cell(row=row_num, column=1).font = Font(bold=True)

# Compter les emails
cursor.execute("SELECT COUNT(*) FROM agents_immobiliers WHERE email IS NOT NULL AND email != '' AND email != 'N/A'")
with_email = cursor.fetchone()[0]

cursor.execute("SELECT COUNT(*) FROM agents_immobiliers WHERE email IS NULL OR email = '' OR email = 'N/A'")
without_email = cursor.fetchone()[0]

row_num += 1
ws.cell(row=row_num, column=1, value=f"Total agents:")
ws.cell(row=row_num, column=2, value=len(agents))

row_num += 1
ws.cell(row=row_num, column=1, value=f"Avec email:")
ws.cell(row=row_num, column=2, value=with_email)

row_num += 1
ws.cell(row=row_num, column=1, value=f"Sans email:")
ws.cell(row=row_num, column=2, value=without_email)

row_num += 1
ws.cell(row=row_num, column=1, value=f"Taux de complétion:")
ws.cell(row=row_num, column=2, value=f"{with_email/len(agents)*100:.1f}%")

# Date de génération
row_num += 2
ws.cell(row=row_num, column=1, value=f"Généré le: {datetime.now().strftime('%Y-%m-%d %H:%M')}")

# Sauvegarder le fichier
output_file = '/root/.openclaw/workspace/agents_immobiliers_idf_final.xlsx'
wb.save(output_file)

print(f"✅ Fichier Excel créé: {output_file}")
print(f"📊 Total agents: {len(agents)}")
print(f"📧 Avec email: {with_email}")
print(f"❌ Sans email: {without_email}")
print(f"📈 Taux: {with_email/len(agents)*100:.1f}%")

conn.close()
